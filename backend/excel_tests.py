import openpyxl

def validate_excel(file_path: str):
    wb = openpyxl.load_workbook(file_path, data_only=False)
    report = {"score": 0, "pass_fail": False, "notes": []}

    # Example 1: Check sheet exists
    if "Sheet1" in wb.sheetnames:
        report["score"] += 3
    else:
        report["notes"].append("Sheet1 missing")
        # try to pick first sheet for further checks
    ws_name = "Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[ws_name]

    # Example 2: Check formula in cell B10
    cell = ws["B10"]
    # openpyxl stores formulas as strings. If file is saved with calculated values only, you might need data_only=True
    if cell.value and isinstance(cell.value, str) and "SUM(" in cell.value:
        # simple heuristic: check if it references A2:A10
        if "SUM(A2:A10)" in cell.value.replace(" ", ""):
            report["score"] += 7
        else:
            report["notes"].append("SUM formula present in B10 but not matching SUM(A2:A10)")
    else:
        report["notes"].append("SUM formula not found in B10")

    # Pass if score >= 8
    report["pass_fail"] = report["score"] >= 8
    return report